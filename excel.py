import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill

from database import get_reports_by_date, get_missed_employees

HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF")


def _style_header(row):
    for cell in row:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")


def _autosize(ws):
    for col in ws.columns:
        values = [str(c.value) for c in col if c.value is not None]
        width = max((len(v) for v in values), default=10)
        ws.column_dimensions[col[0].column_letter].width = width + 4


async def build_excel_report(report_date: str) -> str:
    """
    Berilgan sana bo'yicha hisobotlarni Excel faylga eksport qiladi.
    1-sahifa: barcha topshirilgan hisobotlar (vaqt, holat bilan)
    2-sahifa: hisobot topshirmagan xodimlar
    Qaytaradi: yaratilgan faylning nomi (path)
    """
    reports = await get_reports_by_date(report_date)
    missed = await get_missed_employees(report_date)

    wb = openpyxl.Workbook()

    ws = wb.active
    ws.title = "Hisobotlar"
    ws.append(["№", "F.I.Sh", "Vaqt", "Holat"])
    _style_header(ws[1])
    for i, (user_id, fullname, report_time, status) in enumerate(reports, start=1):
        ws.append([i, fullname, report_time, status])

    ws2 = wb.create_sheet("Topshirmaganlar")
    ws2.append(["№", "F.I.Sh"])
    _style_header(ws2[1])
    for i, (user_id, fullname) in enumerate(missed, start=1):
        ws2.append([i, fullname])

    _autosize(ws)
    _autosize(ws2)

    filename = f"hisobot_{report_date}.xlsx"
    wb.save(filename)
    return filename
